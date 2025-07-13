import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def read_data_file(file_path):
    """Read data from either Excel or CSV file"""
    try:
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path)
        else:  # Assume Excel for .xlsx, .xls
            return pd.read_excel(file_path)
    except Exception as e:
        print(f"Error reading file {file_path}: {str(e)}")
        return None

def validate_columns(df, file_name):
    """Check if required columns exist in the dataframe"""
    required_columns = ['Source NE', 'Destination NE', 'Source Port', 'Destination Port']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"Error in {file_name}: Missing columns - {', '.join(missing_columns)}")
        return False
    return True

def compare_data_files(file1_path, file2_path):
    """Compare two data files and generate a comparison report"""
    # Read both files
    sheet1 = read_data_file(file1_path)
    sheet2 = read_data_file(file2_path)
    
    if sheet1 is None or sheet2 is None:
        return None
    
    # Validate columns
    if not validate_columns(sheet1, "File 1") or not validate_columns(sheet2, "File 2"):
        return None
    
    # Merge for comparison
    merged = pd.merge(
        sheet1, 
        sheet2, 
        on=['Source NE', 'Destination NE'], 
        how='outer', 
        indicator='NE_Match',
        suffixes=('_File1', '_File2')
    
    # Convert match indicator
    merged['NE_Match'] = merged['NE_Match'].map({
        'left_only': 'Mismatched (Only in File1)',
        'right_only': 'Mismatched (Only in File2)',
        'both': 'Matched'
    })
    
    # Compare ports
    def compare_ports(row):
        if row['NE_Match'] != 'Matched':
            return 'N/A'
        
        source_match = str(row['Source Port_File1']) == str(row['Source Port_File2'])
        dest_match = str(row['Destination Port_File1']) == str(row['Destination Port_File2'])
        
        if source_match and dest_match:
            return 'Ports Matched'
        else:
            mismatches = []
            if not source_match:
                mismatches.append(f"Source Port File2: {row['Source Port_File2']}")
            if not dest_match:
                mismatches.append(f"Destination Port File2: {row['Destination Port_File2']}")
            return ', '.join(mismatches)
    
    merged['Port_Comparison'] = merged.apply(compare_ports, axis=1)
    
    # Reorder columns for better readability
    column_order = [
        'Source NE', 'Destination NE',
        'Source Port_File1', 'Source Port_File2',
        'Destination Port_File1', 'Destination Port_File2',
        'NE_Match', 'Port_Comparison'
    ]
    
    # Add any remaining columns
    other_columns = [col for col in merged.columns if col not in column_order]
    merged = merged[column_order + other_columns]
    
    return merged

def save_comparison_result(merged_df, original_path):
    """Save the comparison results to an Excel file with formatting"""
    if merged_df is None:
        return None
    
    output_dir = os.path.dirname(original_path)
    base_name = os.path.splitext(os.path.basename(original_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}_comparison_result.xlsx")
    
    try:
        # Save to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False, sheet_name='Comparison Results')
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets['Comparison Results']
            
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            
            for row in range(2, worksheet.max_row + 1):
                ne_match = worksheet.cell(row=row, column=merged_df.columns.get_loc('NE_Match') + 1).value
                port_match = worksheet.cell(row=row, column=merged_df.columns.get_loc('Port_Comparison') + 1).value
                
                if 'Mismatched' in str(ne_match):
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = red_fill
                elif port_match != 'Ports Matched' and port_match != 'N/A':
                    for col in range(1, worksheet.max_column + 1):
                        header = worksheet.cell(row=1, column=col).value
                        if header and '_File2' in str(header):
                            worksheet.cell(row=row, column=col).fill = red_fill
        
        return output_path
    
    except Exception as e:
        print(f"Error saving results: {str(e)}")
        return None

def main():
    print("Network Element Comparison Tool")
    print("=" * 40)
    
    # Get file paths from user
    file1_path = input("Enter path to first file (File1): ").strip()
    file2_path = input("Enter path to second file (File2): ").strip()
    
    # Validate paths
    if not os.path.exists(file1_path):
        print(f"Error: File not found - {file1_path}")
        return
    if not os.path.exists(file2_path):
        print(f"Error: File not found - {file2_path}")
        return
    
    # Compare files
    comparison_result = compare_data_files(file1_path, file2_path)
    
    if comparison_result is not None:
        # Save results
        output_path = save_comparison_result(comparison_result, file1_path)
        if output_path:
            print(f"\nComparison completed successfully!")
            print(f"Results saved to: {output_path}")
        else:
            print("\nFailed to save comparison results.")
    else:
        print("\nComparison failed. Please check the error messages above.")

if __name__ == "__main__":
    main()
